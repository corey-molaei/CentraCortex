from __future__ import annotations

import subprocess
import tempfile
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None

try:
    from docx import Document as DocxDocument
except Exception:  # pragma: no cover
    DocxDocument = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None


MIME_EXTENSION_HINTS: dict[str, str] = {
    "text/plain": "txt",
    "text/markdown": "md",
    "text/csv": "csv",
    "application/csv": "csv",
    "application/json": "json",
    "application/xml": "xml",
    "application/pdf": "pdf",
    "application/msword": "doc",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/vnd.ms-excel": "xls",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
}

SUPPORTED_EXTENSIONS: set[str] = {
    "txt",
    "md",
    "csv",
    "json",
    "xml",
    "pdf",
    "docx",
    "xlsx",
    "xls",
    "doc",
}


@dataclass
class ExtractedTextResult:
    text: str
    extractor: str
    metadata: dict[str, Any] = field(default_factory=dict)
    warning: str | None = None


def resolve_extension(*, filename: str, mime_type: str | None = None) -> str:
    extension = Path(filename).suffix.lower().lstrip(".")
    if extension:
        return extension
    mime = (mime_type or "").lower().strip()
    return MIME_EXTENSION_HINTS.get(mime, "")


def can_extract_text(*, filename: str, mime_type: str | None = None) -> bool:
    extension = resolve_extension(filename=filename, mime_type=mime_type)
    return extension in SUPPORTED_EXTENSIONS


def extract_text_from_file_bytes(*, filename: str, content: bytes, mime_type: str | None = None) -> ExtractedTextResult:
    extension = resolve_extension(filename=filename, mime_type=mime_type)
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported extension: {extension or 'unknown'}")

    if extension in {"txt", "md", "csv", "json", "xml"}:
        return ExtractedTextResult(
            text=content.decode("utf-8", errors="ignore"),
            extractor="utf8",
            metadata={"extension": extension},
        )

    if extension == "pdf":
        if PdfReader is None:
            raise ValueError("PDF support requires optional dependency 'pypdf'")
        reader = PdfReader(BytesIO(content))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
        return ExtractedTextResult(
            text=text,
            extractor="pypdf",
            metadata={"extension": extension, "pages": len(reader.pages)},
        )

    if extension == "docx":
        if DocxDocument is None:
            raise ValueError("DOCX support requires optional dependency 'python-docx'")
        doc = DocxDocument(BytesIO(content))
        paragraphs = [paragraph.text for paragraph in doc.paragraphs]
        return ExtractedTextResult(
            text="\n".join(paragraphs),
            extractor="python-docx",
            metadata={"extension": extension, "paragraph_count": len(paragraphs)},
        )

    if extension == "xlsx":
        return _extract_xlsx(content)

    if extension == "xls":
        return _extract_xls(content)

    if extension == "doc":
        return _extract_doc(content)

    raise ValueError(f"Unsupported extension: {extension}")


def _extract_xlsx(content: bytes) -> ExtractedTextResult:
    if load_workbook is None:
        raise ValueError("XLSX support requires optional dependency 'openpyxl'")

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    lines: list[str] = []
    sheet_rows: dict[str, int] = {}

    for sheet in workbook.worksheets:
        non_empty_rows = 0
        lines.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            row_text = ", ".join(values).strip()
            if row_text:
                lines.append(row_text)
                non_empty_rows += 1
        sheet_rows[sheet.title] = non_empty_rows

    workbook.close()
    return ExtractedTextResult(
        text="\n".join(lines),
        extractor="openpyxl",
        metadata={
            "extension": "xlsx",
            "sheet_names": list(sheet_rows.keys()),
            "sheet_row_counts": sheet_rows,
            "sheet_count": len(sheet_rows),
            "row_count_total": sum(sheet_rows.values()),
        },
    )


def _extract_xls(content: bytes) -> ExtractedTextResult:
    if xlrd is None:
        raise ValueError("XLS support requires optional dependency 'xlrd'")

    workbook = xlrd.open_workbook(file_contents=content)
    lines: list[str] = []
    sheet_rows: dict[str, int] = {}

    for sheet in workbook.sheets():
        non_empty_rows = 0
        lines.append(f"Sheet: {sheet.name}")
        for row_index in range(sheet.nrows):
            row_values = ["" if value is None else str(value) for value in sheet.row_values(row_index)]
            row_text = ", ".join(row_values).strip()
            if row_text:
                lines.append(row_text)
                non_empty_rows += 1
        sheet_rows[sheet.name] = non_empty_rows

    return ExtractedTextResult(
        text="\n".join(lines),
        extractor="xlrd",
        metadata={
            "extension": "xls",
            "sheet_names": list(sheet_rows.keys()),
            "sheet_row_counts": sheet_rows,
            "sheet_count": len(sheet_rows),
            "row_count_total": sum(sheet_rows.values()),
        },
    )


def _extract_doc(content: bytes) -> ExtractedTextResult:
    temp_path: str | None = None
    try:
        with tempfile.NamedTemporaryFile(prefix="doc_parse_", suffix=".doc", delete=False) as temp_file:
            temp_file.write(content)
            temp_path = temp_file.name

        try:
            proc = subprocess.run(
                ["antiword", temp_path],
                capture_output=True,
                text=True,
                check=False,
                timeout=30,
            )
        except FileNotFoundError as exc:
            raise ValueError("DOC support requires system dependency 'antiword'") from exc

        if proc.returncode != 0:
            detail = (proc.stderr or proc.stdout or "antiword parse failed").strip()
            raise ValueError(f"DOC parsing failed: {detail}")

        return ExtractedTextResult(
            text=(proc.stdout or "").strip(),
            extractor="antiword",
            metadata={"extension": "doc"},
        )
    finally:
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass
